import openpyxl
from unidecode import unidecode
import csv

excel_file = openpyxl.load_workbook('database.xlsx')
spreadsheet = excel_file['Planilha1']

all_users = []
header = ('Nome de usuário,Nome,Sobrenome,Nome de exibição,Cargo,Departamento,Número comercial,Telefone comercial,'
          'Celular,Fax,Endereço de email alternativo,Endereço,Cidade,Estado ou província,CEP,País ou região')


def username_create(name):
    name = unidecode(name).lower().split()
    email = '.'.join(name[::len(name) - 1]) + '@frexco.com.br'
    return email


def name_create(name):
    name = name.split()
    first_name = name[0]
    last_name = ' '.join((name[1::]))
    exibithion_name = ' '.join(name)

    return first_name, last_name, exibithion_name


def generate_csv_file(users):
    with open('office365.csv', 'w', encoding='utf-8') as file:
        file.write(header + '\n')
    for user in users:
        with open('office365.csv', 'a', newline='', encoding='utf-8') as file:
            csv_writer = csv.writer(file)
            row_data = list(user.values()) + [''] * 9
            csv_writer.writerow(row_data)
            print(f'Adicionado usuário ao arquivo CSV: {user['exibithion_name']}')


def extract_excel_data(spreadsheet_name):
    username = all_names = position = department = None
    for line in spreadsheet_name.iter_rows(min_row=2, max_row=spreadsheet_name.max_row, min_col=1, max_col=3):
        if any(cell.value is None for cell in line):
            continue
        user = {}
        for cell in line:
            if cell.column == 1:  # criar alias de e-mail
                cell.value = str(cell.value)
                username = username_create(cell.value)
                all_names = name_create(cell.value)

            elif cell.column == 2:
                position = cell.value

            elif cell.column == 3:
                department = cell.value

        user['username'] = username
        user['first_name'] = all_names[0]
        user['last_name'] = all_names[1]
        user['exibithion_name'] = all_names[2]
        user['position'] = position
        user['department'] = department

        all_users.append(user)


extract_excel_data(spreadsheet)
generate_csv_file(all_users)
excel_file.close()
